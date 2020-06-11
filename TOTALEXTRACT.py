import pandas as pd
import os
import math
import numpy as np
from openpyxl import load_workbook
from itertools import islice
from openpyxl.utils import get_column_letter
from scipy import sparse
from scipy.sparse import csgraph
from scipy.special import softmax
def safely_execute(func):
    def func_wrapper(*args, **kwargs):
        try:
           return func(*args, **kwargs)
        except Exception as e:
            print(e) # or write to log
            return None
    return func_wrapper
@safely_execute
def get_x_info(model):
    print(url)

for  x in range(3,1,-1):
    print("hi",x)
def MIlaplacian(adjm):
    #laplacian: diagonal entries are all 1
    #connected nodes have 1/degree else 0
    degree_list=adjm.sum(axis=1).astype(np.float32)
    #sum up degree for each node, possibility of disconnected node
    degree_list=np.where(degree_list<=0,-1,degree_list)
    coeff=-np.where(degree_list<0,0,1/degree_list)
    for row in range(adjm.shape[0]):
        adjm[row,:]*=coeff[row]
    np.fill_diagonal(adjm,1)
    return adjm





fexcel=os.path.join("D:/USHUR/PythonProg/DATAREP/IL-EXTRACTION/",'IL4.xlsx')
#cells can have formulas
wb = load_workbook(fexcel)
#cells have only values
wbd=load_workbook(fexcel,data_only=True)
#print(wb.sheetnames)
#print(wb.active)
#to change active  worksheet
#wb.active=0
#print(wb.active)
fsheet=wb.active
dsheet=wbd.active
#get row id and col id and value given row, col index
print(dsheet.min_row)
print(dsheet.max_row)
print(dsheet.min_column)
print(dsheet.max_column)

print(dsheet.cell(1,1).row)
print(dsheet.cell(1,1).column)
print(dsheet.cell(1,1).value)
print(dsheet.cell(1,1).font.bold)
if dsheet.cell(1,1).value==' ':
    print("null string")

#get row id and col id and value given row, column as string
#print(dsheet.cell('D4').row)
#print(dsheet.cell('D4').column)
#print(dsheet.cell('D4').value)
#xy = coordinate_from_string('A4') # returns ('A',4)
#col = column_index_from_string(xy[0]) # returns 1

toi_dict={}

edgedict={}
#node_attr=(word_key,distance,type)
#if node is nULL, no edge
node_attributes=[]
doc_adjlist={}
def sweep_doc(doc_instance,for_instance):
    tag='None'
    val='None'
    tty='None'
    for ic, col_cells in enumerate(islice(doc_instance.columns, doc_instance.max_column)):
        for ir, row in enumerate(islice(col_cells, doc_instance.max_row)):
            if row.value!='None':
                if type(row.value)==str and row.value.isspace():
                    val='None'
                else:
                    val=str(row.value).replace('-','')
                    tty=str(type(row.value))[8:-2]
                    #tag only words as keys if bold font
                    if doc_instance.cell(ir + 1, ic + 1).font.bold and tty=='str':
                        tag = 'KEY'
                    else:
                        tag = 'VALUE'
                    if type(for_instance.cell(ir + 1, ic + 1).value) == str and for_instance.cell(ir + 1, ic + 1).value[
                    0] == '=':
                        tty='formula'
                        if for_instance.cell(ir + 1, ic + 1).value[0:4] == '=SUM':
                            val = 'SUM'
                        else:
                            val = 'EQUALS'
            #print("col and row ", ic, ir)
            #if type(row.value) == int or type(row.value)==float:

            #print(str(ir) +'-'+ str(ic) + '-'+str(row.value))
            yield str(ir) +'-'+ str(ic) +'-'+ val +'-'+ tty +'-'+ tag




def NN_edges(r,c,rowmin, colmin, nodelist,augnodelist):
    e,w,s,n,tb,lb=-1,-1,-1,-1,-1,-1
    ei, wi, si, ni, tbi, lbi = -1, -1, -1, -1, -1, -1
    ec, wc, sc, nc, tbc, lbc = -1, -1, -1, -1, -1, -1
    et, wt, st, nt, tbt, lbt = -1, -1, -1, -1, -1, -1
    ev, wv, sv, nv, tbv, lbv = -1, -1, -1, -1, -1, -1
    ekv, wkv, skv, nkv, tbkv, lbkv = -1, -1, -1, -1, -1, -1
    NNE=[]
    if (r,c-1) in nodelist:
        e= 1.0 #(r,c)
        ei=nodelist.index((r,c-1))
        ec=float(str(r) + '.'+ str(c-1))
        et=augnodelist[nodelist.index((r,c-1))].split('-')[3]
        ev = augnodelist[nodelist.index((r, c - 1))].split('-')[2]
        ekv = augnodelist[nodelist.index((r, c - 1))].split('-')[4]
    if  (r,c+1) in nodelist:
        w=1.180 #(r,c+1)
        wi = nodelist.index((r, c +1))
        wc = float(str(r) + '.' + str(c + 1))
        wt = augnodelist[nodelist.index((r, c +1))].split('-')[3]
        wv = augnodelist[nodelist.index((r, c + 1))].split('-')[2]
        wkv = augnodelist[nodelist.index((r, c + 1))].split('-')[4]
    if  (r-1,c) in nodelist:
        n=1.90# (r-1,c)
        ni = nodelist.index((r-1, c ))
        nc = float(str(r-1) + '.' + str(c))
        nt = augnodelist[nodelist.index((r-1, c))].split('-')[3]
        nv = augnodelist[nodelist.index((r - 1, c))].split('-')[2]
        nkv = augnodelist[nodelist.index((r - 1, c))].split('-')[4]
    if  (r+1,c) in nodelist:
        s=1.270 #(r+1,c)
        si = nodelist.index((r+1, c))
        sc = float(str(r+1) + '.' + str(c))
        st = augnodelist[nodelist.index((r+1, c))].split('-')[3]
        sv = augnodelist[nodelist.index((r + 1, c))].split('-')[2]
        skv = augnodelist[nodelist.index((r + 1, c))].split('-')[4]
    if r> rowmin:
        kr=rowmin
        for keyrow in range(r,rowmin,-1):
            if (keyrow-1,c) in nodelist:
                if augnodelist[nodelist.index((keyrow-1, c))].split('-')[4]=='KEY':
                    kr=keyrow-1
                    break
        if (kr,c) in nodelist:
            tb=float(str(kr) + '.' +'90')#(0,c)
            tbi = nodelist.index((kr, c))
            tbc = float(str(kr) + '.' + str(c))
            tbt = augnodelist[nodelist.index((kr, c))].split('-')[3]
            tbv = augnodelist[nodelist.index((kr, c))].split('-')[2]
            tbkv = augnodelist[nodelist.index((kr, c))].split('-')[4]

    if c>colmin:
        kc=colmin
        for keycol in range(c,colmin, -1):
            if (r,keycol-1) in nodelist:
                if augnodelist[nodelist.index((r,keycol-1))].split('-')[4]=='KEY':
                    kc=keycol-1
                    break
        if (r,kc) in nodelist:
            lb=float(str(kc) + '.' + '180')#(r,0)
            lbi = nodelist.index((r, kc))
            lbc = float(str(r) + '.' + str(kc))
            lbt = augnodelist[nodelist.index((r, kc))].split('-')[3]
            lbv = augnodelist[nodelist.index((r, kc))].split('-')[2]
            lbkv = augnodelist[nodelist.index((r, kc))].split('-')[4]
    NNE.append([ei,ec,e,ev,et,ekv])
    NNE.append([wi,wc,w,wv,wt,wkv])
    NNE.append([si,sc,s,sv,st,skv])
    NNE.append([ni,nc,n,nv,nt,nkv])
    NNE.append([tbi,tbc,tb,tbv,tbt,tbkv])
    NNE.append([lbi,lbc,lb,lbv,lbt,lbkv])
    return NNE



labeled_node_list=[ x for x in list(sweep_doc(dsheet,fsheet)) if x.split('-')[2]!='None']
#token_loc_list=[x for x in token_loc_list if x.split('-')[2]!='None']
print(len(labeled_node_list))
print("non null tokens", labeled_node_list)
set_of_nodes=[(int(x.split('-')[0]),int(x.split('-')[1])) for x in labeled_node_list]
print(set_of_nodes)
l1,l2=zip(*set_of_nodes)
rowmin,rowmax,colmin,colmax=min(l1),max(l1),min(l2),max(l2)
print(rowmin,rowmax,colmin,colmax)


def get_node_features(nodeelement,rowmin,colmin,set_of_nodes,token_desc_list):
    ni=set_of_nodes.index(nodeelement)
    nc=float(str(nodeelement[0]) + '.'+ str(nodeelement[1]))
    nd=0
    nv=token_desc_list[ni].split('-')[2]
    nt = token_desc_list[ni].split('-')[3]
    ny = token_desc_list[ni].split('-')[4]
    ADJ=[]
    ADJ.append([ni, nc, nd,nv,nt,ny])
    ADJ.append(NN_edges(nodeelement[0], nodeelement[1], rowmin,colmin,set_of_nodes, token_desc_list))
    return ADJ

#adjlist=[get_node_features(x,set_of_nodes,labeled_node_list) for x in set_of_nodes]
#print(adjlist)
fadjlist=[get_node_features(x,rowmin,colmin,set_of_nodes,labeled_node_list) for x in set_of_nodes]
print(fadjlist)

def prune(adjrow):
    ADJ=[]
    ADJ.append(adjrow[0])
    #node features as is
    edgelist=[]
    #prune edges that satisfy a given condition (V,V)
    #print("adj",ADJ)
    #print(adjrow[1])
    for x in adjrow[1]:
        if x[5]=='VALUE':
            if (adjrow[0][4]=='int' or adjrow[0][4]=='float') and (x[4]=='int' or x[4]=='float'):
                edgelist.append([-1, -1, -1, -1, -1, -1])
            else:
                if (adjrow[0][4] == 'str' and x[4] == 'str'):
                    edgelist.append([-1, -1, -1, -1, -1, -1])
                else:
                    edgelist.append(x)

            #print("edge removed",edgelist)
        else:
            edgelist.append(x)
            #print("edge retained",edgelist)
    ADJ.append(edgelist)
    #print("return ADJ",ADJ)

    return ADJ



pruneadjlist=[prune(x) if x[0][5]=='VALUE' else x for x in fadjlist]
print(pruneadjlist)

#print("35",pruneadjlist[35])
'''
#print(adjlist[0][5][0])
print(len(pruneadjlist))
adj_matrix=np.zeros((len(pruneadjlist),len(pruneadjlist)))
degree_list=np.zeros(len(pruneadjlist))
for row, column in enumerate(pruneadjlist):
     for edgeno, edge in enumerate(column[1]):
         if edge[0]!=-1:
             adj_matrix[row,edge[0]]=1

adj_matrix=np.array(adj_matrix,dtype=float)
print(adj_matrix[0])
ladj_matrix=MIlaplacian(adj_matrix)
print(ladj_matrix)

e_f=0
GLOVE_DIR="d:/USHUR/PythonProg/DATAREP/Embeddings/glove.6B"
embeddings_index = {}
f = open(os.path.join(GLOVE_DIR, 'glove.6B.300d.txt'),encoding='utf8')
for line in f:
    values = line.split()
    word = values[0]
    coefs = np.asarray(values[1:], dtype='float32')
    embeddings_index[word] = coefs
f.close()
print('Found %s word vectors.' % len(embeddings_index))
embedding_dim=2100
EM=np.zeros((0,embedding_dim))
for row in pruneadjlist:
    eb=np.zeros(300)
    edge_emb=np.zeros((6,300))
    if row[0][5] == 'VALUE':
        if row[0][4] == 'str':
            row[0][3] = 'word'
        else:
            if row[0][4] == 'int' or row[0][4]=='float':
                row[0][3]='number'
    node_emb = np.mean([ eb if embeddings_index.get(x) is None else embeddings_index.get(x) for x in row[0]],axis=0)
    for eno, edge in enumerate(row[1]):
        if edge[5]=='VALUE':
            if edge[4]=='str':
                edge[3]='word'
            else:
                if edge[4]=='int' or edge[4]=='float':
                    edge[3]='number'
        edge_emb[eno] = np.mean([eb if embeddings_index.get(x) is None else embeddings_index.get(x) for x in edge],axis=0)
    row_emb=np.hstack((node_emb,edge_emb[0],edge_emb[1],edge_emb[2],edge_emb[3],edge_emb[4],edge_emb[5]))
    #print(row_emb.shape)
    EM=np.append(EM,row_emb.reshape(1,-1),axis=0)
print(" EM shape",EM.shape)
LEM=np.matmul(ladj_matrix,EM)
print("LEM shape",LEM.shape)
'''
foXC1=os.path.join("d:/USHUR/PythonProg/DATAREP/IL-EXTRACTION","XC1.dat")
# save matrix for future use if needed
#LEM.dump(foXC5)

embedding_matrix1=np.load(foXC1,allow_pickle=True)

foXC2=os.path.join("d:/USHUR/PythonProg/DATAREP/IL-EXTRACTION","XC2.dat")
# save matrix for future use if needed
#LEM.dump(foXC1)
embedding_matrix2=np.load(foXC2,allow_pickle=True)

foXC3=os.path.join("d:/USHUR/PythonProg/DATAREP/IL-EXTRACTION","XC3.dat")
# save matrix for future use if needed
#LEM.dump(foXC1)
embedding_matrix3=np.load(foXC3,allow_pickle=True)

foXC4=os.path.join("d:/USHUR/PythonProg/DATAREP/IL-EXTRACTION","XC4.dat")
# save matrix for future use if needed
#LEM.dump(foXC1)
embedding_matrix4=np.load(foXC4,allow_pickle=True)
#if embedding matrix is not already computed
#embedding_matrix2=np.load(foXC2,allow_pickle=True)
foXC5=os.path.join("d:/USHUR/PythonProg/DATAREP/IL-EXTRACTION","XC5.dat")
embedding_matrix5=np.load(foXC5,allow_pickle=True)

total_node=embedding_matrix1[58]
'''
print(total_node.shape,np.transpose(total_node).shape)
print(np.matmul(embedding_matrix1,np.transpose(total_node)))
print(np.argmax(np.matmul(embedding_matrix1,np.transpose(total_node))))
print("output shape",np.matmul(embedding_matrix1,np.transpose(total_node)).shape)
exit()

print(np.matmul(embedding_matrix2,np.transpose(total_node)))
print(np.argmax(np.matmul(embedding_matrix2,np.transpose(total_node))))
print("output shape",np.matmul(embedding_matrix2,np.transpose(total_node)).shape)

print(np.matmul(embedding_matrix3,np.transpose(total_node)))
print(np.argmax(softmax(np.matmul(embedding_matrix3,np.transpose(total_node)))))
print("output shape",np.matmul(embedding_matrix3,np.transpose(total_node)).shape)
'''
print(np.matmul(embedding_matrix4,np.transpose(total_node)))
print(np.argmax(softmax(np.matmul(embedding_matrix4,np.transpose(total_node)))))
print("output shape", np.matmul(embedding_matrix4,np.transpose(total_node)).shape)

# code for glove embedding ends here
exit()
#    for edge in row[1]:
#        print("hello")


def sweep_doc_for_type(doc_instance):
    index=0
    for ic, col_cells in enumerate(islice(doc_instance.columns, doc_instance.max_column)):
        for ir, row in enumerate(islice(col_cells, doc_instance.max_row)):
            #print("col and row ", ic, ir)
            #if type(row.value) == int or type(row.value)==float:
            print(str(ic) +'-'+ str(ir) + '-'+str(type(row.value)))
            yield index, str(type(row.value))[8:-2]
            index=index+1

def sweep_doc_for_specialtags(doc_instance):
    index=0
    for ic, col_cells in enumerate(islice(doc_instance.columns, doc_instance.max_column)):
        for ir, row in enumerate(islice(col_cells, doc_instance.max_row)):
            if type(row.value)==str and row.value[0]=='=':
            #print("col and row ", ic, ir)
            #if type(row.value) == int or type(row.value)==float:
            #print(str(ic) +'-'+ str(ir) + '-'+str(row.value))
                yield index, 'formula'+str(row.value[0:4])
            index=index+1

def check_for_formulas(node,doc_instance):
    r=int(node.split('-')[0])
    c=int(node.split('-')[1])
    print(r,c)
    if type(doc_instance.cell(r+1,c+1).value)==str and doc_instance.cell(r+1,c+1).value[0]=='=':
        if doc_instance.cell(r+1,c+1).value[0:4]=='=SUM':
            t='SUM'
        else:
            t='EQUALS'
        #print('formula found',doc_instance.cell(r+1,c+1).value)
        return str(r) + '-' + str(c) + '-' + t + '-' + 'formula'

    else:
        return node
def add_tags(node,doc_instance):
    r=int(node.split('-')[0])
    c=int(node.split('-')[1])
    #print(r,c)
    if doc_instance.cell(r+1,c+1).font.bold:
        t='KEY'
    else:
        t='VALUE'
    return  node + '-' +t


exit()
#sadj_matrix=sparse.csr_matrix(adj_matrix)
#print(sadj_matrix)
#ladj_matrix=csgraph.laplacian(sadj_matrix,normed=False)
#print(ladj_matrix)

exit()



if (1,0) in set_of_nodes:
    print('HI',set_of_nodes.index((1,0)))
exit()
for x in token_loc_list:
    check_for_formulas(x,fsheet)

exit()
token_type_dict=dict(sweep_doc_for_type(dsheet))
print(token_type_dict)
exit()
token_type_dict.update(sweep_doc_for_specialtags(fsheet))
print(token_loc_list)
print(token_type_dict)

exit()

def NN_type(r,c,ws):
    e,w,s,n,tb,lb=0,0,0,0,0,0
    if c>1 and isinstance(ws.cell(r,c-1).value,str):
        e= ws.cell(r,c-1).value
    if  c<ws.max_column and isinstance(ws.cell(r,c+1).value,str):
        w=ws.cell(r,c+1).value
    if  r>1 and isinstance(ws.cell(r-1,c).value,str):
        n=ws.cell(r-1,c).value
    if  r<ws.max_row and isinstance(ws.cell(r+1,c).value,str):
        s=ws.cell(r+1,c).value
    if isinstance(ws.cell(1,c).value,str):
        tb=ws.cell(1,c).value
    if isinstance(ws.cell(r,1).value, str):
        lb=ws.cell(r, 1).value
    return e,w,s,n,tb,lb

TOPKDICT={}
TOPK=[]
#way to extract cell indices as integers
for ic, col_cells in enumerate(islice(fsheet.columns,fsheet.max_column)):
    #iterates all rows for each column
    for ir, row in enumerate(islice(col_cells,fsheet.max_row)):
        #iterates within each column all rows
        if type(row.value)==str and row.value[1:4]=='SUM':
            #print('%s%s: cell.value= %s' % (get_column_letter(ic+1),ir+1,row.value) )
            #print("neighborhood", NN_type(ir+1,ic+1,fsheet))
            TOPK.append(dsheet.cell(ir+1,ic+1).value)
if TOPK:
    for ic, col_cells in enumerate(islice(dsheet.columns, dsheet.max_column)):
        for ir, row in enumerate(islice(col_cells, dsheet.max_row)):
            print("col and row ", ic, ir)
            if type(row.value) == int or type(row.value)==float:

                print("neighborhood", NN_type(ir + 1, ic + 1, dsheet))
                print('%s%s: cell.value= %s' % (get_column_letter(ic+1),ir+1,row.value) )
                TOPKDICT.update({str(ic+1)+ '-'+str(ir+1):dsheet.cell(ir + 1, ic + 1).value})


print(TOPKDICT)

print(max(TOPK))
exit()


